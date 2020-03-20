import csv
import ParserLogic
import openpyxl
from pathlib import Path

def getRatios(dataDirectory):
    companyName = ""
    companyTicker = ""
    #with open(dataDirectory) as csvfile:
    #reader = csv.reader(csvfile)
    totalAssets = 0.0
    totalLiabilities = 0.0
    totalEquity = 0.0
    currentLiabilities = 0.0
    currentAssets = 0.0
    cash = 0.0
    treasuryStock = 0
    totalEquity = 0.0
    revenue = 0.0
    earnings = 0.0
    # Parse input data
    xlsx_file = Path(dataDirectory)
    wb_obj = openpyxl.load_workbook(xlsx_file)
    # Read the first sheet:
    balanceSheetName = ""
    statementOfOperationsName = ""
    coverSheet = wb_obj.active
    for sheetName in wb_obj.sheetnames:
        if(sheetName.lower() == "consolidated balance sheets"):
            balanceSheetName = sheetName
        elif(sheetName.lower() == "consolidated statements of oper"):
            statementOfOperationsName = sheetName
    for row in coverSheet.iter_rows():
        if(row[0].value != None):
            aspect = row[0].value.lower()
            if(aspect == "trading symbol"):
                companyTicker = row[1].value
            elif(aspect == "entity registrant name"):
                companyName = row[1].value
    balanceSheet = wb_obj[balanceSheetName]
    operationsStatement = wb_obj[statementOfOperationsName]
    for row in balanceSheet.iter_rows():
        try:
            aspect = row[0].value.lower()
        except AttributeError:
            aspect = ""

        if(aspect == 'total assets'):
            totalAssets = ParserLogic.getNearestRowValue(row)
        elif("total liabilities" == aspect):
            totalLiabilities = ParserLogic.getNearestRowValue(row)
        elif(aspect == 'total current liabilities'):
            currentLiabilities = ParserLogic.getNearestRowValue(row)
        elif("cash" in aspect):
            cash = ParserLogic.getNearestRowValue(row)
        elif(aspect == 'total current assets'):
            currentAssets = ParserLogic.getNearestRowValue(row)
        elif("treasury stock" in aspect):
            treasuryStock = ParserLogic.getNearestRowValue(row)
        elif(ParserLogic.isTotalStockHoldersEquity(aspect)):
        #elif("Total stockholders’ equity (deficit)" == aspect):
            totalEquity = ParserLogic.getNearestRowValue(row)

    for row in operationsStatement.iter_rows():
        try:
            aspect = row[0].value.lower()
        except AttributeError:
            aspect = ""

        if(aspect == "net revenues" or aspect == "total revenues" or aspect == "operating revenues" or aspect == "net sales"):
            revenue = ParserLogic.getNearestRowValue(row)
        elif(aspect == "net earnings" or aspect == "net loss" or aspect == "net income"):
            earnings = ParserLogic.getNearestRowValue(row)

    # Calculate Ratios
    if(totalEquity == 0 and totalLiabilities != 0):
        totalEquity = totalAssets-totalLiabilities
    if(totalLiabilities == 0 and totalAssets != 0 and totalEquity != 0):
        totalLiabilities = totalAssets - totalEquity 
    #print("Equity: "+str(totalEquity)+"\nLiabilties: "+str(totalLiabilities)+"\nAssets: "+str(totalAssets))
    debt_to_assets = totalLiabilities/totalAssets
    longTermLiabilities = totalLiabilities - currentLiabilities
    longtermDE = longTermLiabilities/totalEquity
    debt_to_equity = totalLiabilities/totalEquity
    
    try:
        cashRatio = cash/currentLiabilities
        current_Ratio = currentAssets/currentLiabilities
    except ZeroDivisionError:
        cashRatio = "No current Liabilities (div by zero)"
        current_Ratio = "No current Liabilities (div by zero)"
    
    if(revenue == 0):
        netPorfitMargin = "Rev was 0"
    else:
        netPorfitMargin = earnings/revenue

    # Print output file
    dumpName = companyTicker + "dump.csv"
    with open(dumpName, 'w', newline='') as dumpFile:
        writer = csv.writer(dumpFile)
        writer.writerow(["Company", companyName])
        writer.writerow(["Ratio", "Value"])
        writer.writerow(["Cash Ratio", str(cashRatio)])
        writer.writerow(["Current Ratio", str(current_Ratio)])
        writer.writerow(["Debt-to-assets Ratio", str(debt_to_assets)])
        writer.writerow(["Debt-to-equity Ratio", str(debt_to_equity)])
        writer.writerow(["Longterm DE", str(longtermDE)])
        writer.writerow(["Treasury Stock", str(treasuryStock)])
        writer.writerow(["Net Profit margin", str(netPorfitMargin)])
    return dumpName